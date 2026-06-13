# ─────────────────────────────────────────────────────────────────────────────
#  Job Tracker — Sync Script
#
#  What this does:
#    Every 5 minutes, pulls your live Google Sheet (JobHunt 2026 tab) and
#    writes all changes directly into your original Excel file:
#      D:\01. SelfBuilding [Personal]\01. Life\01. Job Hunting\
#        04. Tracker And Analysis\JobTracker_Organized.xlsx
#
#  It updates:
#    • New rows added by the Chrome extension (via Google Sheet)
#    • Status changes detected by Gmail (Rejected / Interviewing / Offer)
#    • "Gmail Synced On" timestamp column
#
#  Runs silently in the background. Set up via install_startup.bat.
# ─────────────────────────────────────────────────────────────────────────────

import pandas as pd
import os
import time
import glob
import traceback
from datetime import datetime
from openpyxl import load_workbook
# Q1 fix: removed unused `import sys`
# Q2 fix: removed unused `from openpyxl.utils.dataframe import dataframe_to_rows`

# ── Paths ─────────────────────────────────────────────────────────────────────

EXCEL_FILE  = r"D:\01. SelfBuilding [Personal]\01. Life\01. Job Hunting\04. Tracker And Analysis\JobTracker_Organized.xlsx"
EXCEL_SHEET = "JobHunt 2026"
DOWNLOADS   = os.path.join(os.path.expanduser("~"), "Downloads")
CSV_PATTERN = os.path.join(DOWNLOADS, "job_tracker*.csv")

# ── Google Drive D: sync path ──────────────────────────────────────────────────
# Because your D: drive is already synced to Google Drive, the Apps Script
# writes a CSV (job_tracker_gs_export.csv) directly into the same Drive folder.
# Google Drive Desktop syncs it here automatically — no URL or CSV publishing needed.
#
# This path must match where your Excel file lives (same folder):
DRIVE_EXPORT_CSV = r"D:\01. SelfBuilding [Personal]\01. Life\01. Job Hunting\04. Tracker And Analysis\job_tracker_gs_export.csv"

# Legacy: leave blank — no longer needed when D: drive syncs to Google Drive
GOOGLE_SHEET_CSV_URL = ""

# ── Timing ────────────────────────────────────────────────────────────────────
POLL_INTERVAL         = 10    # seconds between CSV export checks
SHEETS_SYNC_INTERVAL  = 300   # seconds between Google Sheet pulls (5 min)

# ── Column names (must match your Excel headers exactly) ──────────────────────
EXCEL_COLS = [
    "Date of Apply", "Organization", "Location", "Role",
    "Exp. Required", "Submission Status", "Portal",
    "Salary", "URL", "Referred by", "Result", "Gmail Synced On"
]

# Columns from the extension CSV → Excel column names
CSV_RENAME = {
    "Date of Apply":      "Date of Apply",
    "Organization":       "Organization",
    "Salary":             "Salary",
    "Location":           "Location",
    "Role":               "Role",
    "Year of experience": "Exp. Required",
    "Submission Status":  "Submission Status",
    "Portal":             "Portal",
    "URL":                "URL",
    "Referred by":        "Referred by",
    "Result":             "Result",
}

# ── Logging ───────────────────────────────────────────────────────────────────

LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sync_log.txt")

LOG_MAX_LINES = 5000  # Q4 fix: rotate log after this many lines to prevent unbounded growth

def log(msg):
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line, flush=True)
    try:
        # Rotate: keep only the last LOG_MAX_LINES lines
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, "r", encoding="utf-8") as f:
                lines = f.readlines()
            if len(lines) >= LOG_MAX_LINES:
                with open(LOG_FILE, "w", encoding="utf-8") as f:
                    f.writelines(lines[-(LOG_MAX_LINES // 2):])  # keep last half
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass

# ── Excel helpers ─────────────────────────────────────────────────────────────

def load_excel():
    if not os.path.exists(EXCEL_FILE):
        log(f"⚠  Excel not found: {EXCEL_FILE}")
        return pd.DataFrame(columns=EXCEL_COLS)
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=EXCEL_SHEET, dtype=str)
        df = df.fillna("")
        # Ensure Gmail Synced On column exists
        if "Gmail Synced On" not in df.columns:
            df["Gmail Synced On"] = ""
        return df
    except Exception as e:
        log(f"✗ Could not read Excel: {e}")
        return pd.DataFrame(columns=EXCEL_COLS)


def save_excel(df):
    """Write the updated DataFrame back into the original Excel, preserving all other sheets."""
    try:
        if not os.path.exists(EXCEL_FILE):
            log(f"✗ Excel file missing — cannot save.")
            return False

        wb = load_workbook(EXCEL_FILE)

        if EXCEL_SHEET not in wb.sheetnames:
            log(f"✗ Sheet '{EXCEL_SHEET}' not found in workbook.")
            return False

        ws = wb[EXCEL_SHEET]

        # Get existing headers from row 1
        existing_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        # Ensure "Gmail Synced On" column exists
        if "Gmail Synced On" not in existing_headers:
            next_col = ws.max_column + 1
            ws.cell(1, next_col).value = "Gmail Synced On"
            existing_headers.append("Gmail Synced On")

        # Map DataFrame columns to Excel column positions
        col_pos = {h: i + 1 for i, h in enumerate(existing_headers) if h}

        # Clear data rows (keep header, keep formatting)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # Write DataFrame rows
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for col_name, pos in col_pos.items():
                if col_name in df.columns:
                    val = row.get(col_name, "")
                    ws.cell(r_idx, pos).value = str(val) if val not in ("", None) else None

        wb.save(EXCEL_FILE)
        log(f"✓ Excel saved — {len(df)} rows in '{EXCEL_SHEET}'.")
        return True

    except PermissionError:
        log("✗ Excel is open and locked. Close the file and the sync will retry.")
        return False
    except Exception as e:
        log(f"✗ Save error: {e}")
        traceback.print_exc()
        return False

# ── CSV merge (extension exports) ─────────────────────────────────────────────

def merge_csv_exports(existing_df):
    files = glob.glob(CSV_PATTERN)
    if not files:
        return existing_df, 0

    total_added = 0
    for fpath in files:
        if os.path.getsize(fpath) == 0:
            os.remove(fpath)
            continue
        try:
            new_df = pd.read_csv(fpath, dtype=str).fillna("")
            new_df = new_df.rename(columns=CSV_RENAME)
            for c in EXCEL_COLS:
                if c not in new_df.columns:
                    new_df[c] = ""

            existing_urls = set(existing_df["URL"].dropna())
            new_rows = new_df[~new_df["URL"].isin(existing_urls)]

            if len(new_rows) > 0:
                existing_df = pd.concat([existing_df, new_rows[EXCEL_COLS]], ignore_index=True)
                total_added += len(new_rows)
                log(f"  CSV: +{len(new_rows)} row(s) from {os.path.basename(fpath)}")

            os.remove(fpath)
        except Exception as e:
            log(f"  CSV error ({os.path.basename(fpath)}): {e}")

    return existing_df, total_added

# ── Google Drive CSV sync (replaces URL polling) ───────────────────────────────
#
# The Apps Script writes job_tracker_gs_export.csv into the same Google Drive
# folder as your Excel. Since D: drive syncs to Google Drive, that CSV appears
# locally at DRIVE_EXPORT_CSV. We read it, merge it into Excel, then delete it
# so the next write from Apps Script is a clean fresh export.

def sync_from_drive_csv(existing_df):
    """Read the CSV the Apps Script dropped into the Drive folder, merge, delete."""
    if not os.path.exists(DRIVE_EXPORT_CSV):
        return existing_df, 0, 0, False

    file_age = time.time() - os.path.getmtime(DRIVE_EXPORT_CSV)
    if file_age > 3600:
        # Stale file (> 1 hour old) — skip to avoid merging outdated data
        log(f"  Drive CSV: file is {int(file_age/60)} min old — skipping (will use next fresh export).")
        return existing_df, 0, 0, False

    try:
        sheet_df = pd.read_csv(DRIVE_EXPORT_CSV, dtype=str).fillna("")
        log(f"  Drive CSV: read {len(sheet_df)} rows from {os.path.basename(DRIVE_EXPORT_CSV)}")


        added   = 0
        updated = 0

        # Update existing rows: Result, Submission Status, Gmail Synced On
        if "URL" in existing_df.columns and "URL" in sheet_df.columns:
            sheet_by_url = sheet_df.set_index("URL")

            for idx, row in existing_df.iterrows():
                url = str(row.get("URL", "")).strip()
                if not url or url not in sheet_by_url.index:
                    continue
                sheet_row = sheet_by_url.loc[url]
                for col in ["Result", "Submission Status", "Gmail Synced On"]:
                    if col not in sheet_row.index:
                        continue
                    new_val = str(sheet_row[col] or "").strip()
                    old_val = str(row.get(col, "") or "").strip()
                    if new_val and new_val != old_val:
                        existing_df.at[idx, col] = new_val
                        updated += 1

        # Add new rows from the Sheet that aren't yet in Excel
        if "URL" in existing_df.columns and "URL" in sheet_df.columns:
            existing_urls = set(existing_df["URL"].dropna().astype(str))
            to_add = sheet_df[~sheet_df["URL"].astype(str).isin(existing_urls)].copy()
            if len(to_add) > 0:
                for c in EXCEL_COLS:
                    if c not in to_add.columns:
                        to_add[c] = ""
                cols_to_use = [c for c in EXCEL_COLS if c in to_add.columns]
                existing_df = pd.concat([existing_df, to_add[cols_to_use]], ignore_index=True)
                added += len(to_add)

        if updated: log(f"  Drive CSV: {updated} field(s) updated (Result / Status / Synced).")
        if added:   log(f"  Drive CSV: +{added} new row(s) added from Google Sheets.")

        # C3 fix: do NOT delete the CSV here.
        # Deletion is deferred to main() AFTER save_excel() succeeds.
        # If save fails (e.g. Excel is open), we return the flag and main() skips deletion,
        # so the CSV stays and will be retried next poll cycle.
        return existing_df, added, updated, True   # True = pending_delete

    except Exception as e:
        log(f"  Drive CSV sync error: {e}")
        traceback.print_exc()
        return existing_df, 0, 0, False   # False = do not delete CSV (error occurred)


def sync_from_google_sheets(existing_df):
    """Legacy URL-based pull — kept as fallback if GOOGLE_SHEET_CSV_URL is set."""
    if not GOOGLE_SHEET_CSV_URL:
        return existing_df, 0, 0

    try:
        sheet_df = pd.read_csv(GOOGLE_SHEET_CSV_URL, dtype=str).fillna("")
        log(f"  Sheets (URL): pulled {len(sheet_df)} rows.")
        added = updated = 0

        if "URL" in existing_df.columns and "URL" in sheet_df.columns:
            sheet_by_url = sheet_df.set_index("URL")
            for idx, row in existing_df.iterrows():
                url = str(row.get("URL", "")).strip()
                if not url or url not in sheet_by_url.index:
                    continue
                sheet_row = sheet_by_url.loc[url]
                for col in ["Result", "Submission Status", "Gmail Synced On"]:
                    if col not in sheet_row.index:
                        continue
                    new_val = str(sheet_row[col] or "").strip()
                    old_val = str(row.get(col, "") or "").strip()
                    if new_val and new_val != old_val:
                        existing_df.at[idx, col] = new_val
                        updated += 1

            existing_urls = set(existing_df["URL"].dropna().astype(str))
            to_add = sheet_df[~sheet_df["URL"].astype(str).isin(existing_urls)].copy()
            if len(to_add) > 0:
                for c in EXCEL_COLS:
                    if c not in to_add.columns:
                        to_add[c] = ""
                existing_df = pd.concat([existing_df, to_add[[c for c in EXCEL_COLS if c in to_add.columns]]], ignore_index=True)
                added += len(to_add)

        if updated: log(f"  Sheets (URL): {updated} field(s) updated.")
        if added:   log(f"  Sheets (URL): +{added} new row(s) added.")
        return existing_df, added, updated

    except Exception as e:
        log(f"  Sheets URL sync error: {e}")
        return existing_df, 0, 0

# ── Main loop ─────────────────────────────────────────────────────────────────

def main():
    log("=" * 60)
    log("Job Tracker Sync Script started")
    log(f"   Excel      : {EXCEL_FILE}")
    log(f"   Drive CSV  : {DRIVE_EXPORT_CSV}")
    log(f"   Sheets URL : {'configured (legacy)' if GOOGLE_SHEET_CSV_URL else 'not set (using Drive CSV — no URL needed)'}")
    log("=" * 60)
    log("How it works:")
    log("  - Apps Script writes job_tracker_gs_export.csv to Google Drive hourly")
    log("  - Google Drive Desktop syncs it to your D: drive automatically")
    log("  - This script picks it up, merges updates into Excel, then deletes it")

    last_sheets_sync = 0

    while True:
        try:
            # P1 fix: skip load_excel entirely when nothing has arrived.
            # Avoids a redundant Excel read/write on every 10-second poll when idle.
            has_extension_csv = bool(glob.glob(CSV_PATTERN))
            has_drive_csv     = os.path.exists(DRIVE_EXPORT_CSV)
            has_sheets_url    = bool(GOOGLE_SHEET_CSV_URL and
                                     (time.time() - last_sheets_sync >= SHEETS_SYNC_INTERVAL))

            if not (has_extension_csv or has_drive_csv or has_sheets_url):
                time.sleep(POLL_INTERVAL)
                continue

            df = load_excel()
            changed = False
            pending_delete = False  # C3 fix: True when Drive CSV should be deleted post-save

            # 1. Merge any CSV exports dropped by the extension into Downloads
            if has_extension_csv:
                df, csv_added = merge_csv_exports(df)
                if csv_added > 0:
                    changed = True

            # 2. Primary: read the Drive-synced CSV that Apps Script wrote hourly
            if has_drive_csv:
                # C3 fix: 4-tuple return; pending_delete=True means "delete after save"
                df, drive_added, drive_updated, pending_delete = sync_from_drive_csv(df)
                if drive_added + drive_updated > 0:
                    changed = True

            # 3. Fallback: legacy URL-based pull (only if GOOGLE_SHEET_CSV_URL is set)
            if has_sheets_url:
                df, added, updated = sync_from_google_sheets(df)
                if added + updated > 0:
                    changed = True
                last_sheets_sync = time.time()

            if changed:
                saved = save_excel(df)
                # C3 fix: only delete Drive CSV after a confirmed successful Excel save.
                # If Excel is open/locked and save fails, CSV stays and retries next cycle.
                if saved and pending_delete:
                    try:
                        os.remove(DRIVE_EXPORT_CSV)
                        log("  Drive CSV: deleted after successful save.")
                    except Exception as del_err:
                        log(f"  Drive CSV: could not delete -- {del_err}")

        except Exception as e:
            log(f"Unexpected error: {e}")
            traceback.print_exc()

        time.sleep(POLL_INTERVAL)


if __name__ == "__main__":
    main()
